from openpyxl import Workbook, load_workbook
from pyad import aduser

def update_ad():
   wb = load_workbook(input('Please enter the name of the excel file: '))
   ws = wb.active
   ad_property = []

   # Getting the properties to be set
   for col in ws.iter_cols(max_row=1):
      for cell in col:
         if cell.value != None:
            ad_property.append(cell.value)
   #print(ad_property)

   # Run through the name in the spreadsheet and update the properties
   for row in ws.iter_rows(min_row=2):
      try:
         c0, *c1 = row
         if c0.value != None:
            user = aduser.ADUser.from_cn(c0.value)
            for p, v in zip(ad_property[1:], c1):
               print('Updating user "%s", property "%s" to "%s"' % (c0.value, p, v.value))
               user.update_attribute(p, v.value)
      except:
         print('Error updating user "%s"' % (c0.value))
      finally:
         input('Press any key to contimue ...')

if __name__ == '__main__':
    update_ad()